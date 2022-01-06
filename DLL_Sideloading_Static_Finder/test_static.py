from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment  
def write_excel(path_to_pe, list_dll, number):
    try:
        wb = load_workbook("..\report\DLL_Sideloading_static.xlsx")
    except:
        wb = Workbook()
        wb.save("..\report\DLL_Sideloading_static.xlsx")

    ws = wb.worksheets[0]
    start = ws.max_row
    start = start +1
    for dll in list_dll:
        temp_row = [number, path_to_pe, dll]
        ws.append(temp_row)
    #wb.save("report.xlsx")
    end = ws.max_row
    if len(list_dll) == 1:
        wb.save("..\report\DLL_Sideloading_static.xlsx")
        return
    #print("merge")
    #wb.save("test.xlsx")
    #now it's time to merge cell
    sheet = wb.active
    sheet.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    top_left_cell_number = ws.cell(row= start, column=1)
    top_left_cell_number.alignment = Alignment(horizontal="center", vertical="center")
    #
    sheet.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
    top_left_cell_path = ws.cell(row= start, column=2)
    top_left_cell_path.alignment = Alignment(horizontal="center", vertical="center")
    wb.save("report.xlsx")
from procmon_parser import Rule, dump_configuration, load_configuration, ProcmonLogsReader
import win32process as proc
import win32api
import time
import os

def export(path_to_pe):
    pe_name = path_to_pe.split('\\')[-1]
    with open("ProcmonConfiguration.pmc", "rb") as f:
        config = load_configuration(f)
        config["FilterRules"][0] = Rule('Process_Name', 'is', pe_name, 'include')
    pmc_file = pe_name.replace('.','_') + ".pmc"
    with open(pmc_file,"wb") as ff:
        dump_configuration(config, ff)
    szProcmonApp = "Procmon.exe"
    szProcmonStartArgs = "/quiet /AcceptEula /Minimized /LoadConfig " + pmc_file + " /backingfile result.pml"
    ok = True
    global dem
    hProcessProcmon, hThreadProcmon, dwProcessIdProcmon, dwThreadIdProcmon = proc.CreateProcess(
                szProcmonApp, szProcmonStartArgs, None, None, 0,
                proc.CREATE_NO_WINDOW, None, None, proc.STARTUPINFO())
    time.sleep(5)
    try:
        
        hProcess, hThread, dwProcessId, dwThreadId = proc.CreateProcess(
                    path_to_pe, None, None, None, 0,
                    proc.CREATE_NO_WINDOW, None, None, proc.STARTUPINFO())
    except:
        ok =False
    time.sleep(10)

    proc.CreateProcess(
                szProcmonApp, "/quiet /Terminate", None, None, 0,
                proc.CREATE_NO_WINDOW, None, None, proc.STARTUPINFO())
    #now it's time to read the result
    time.sleep(20)
    if ok == False:
        return list()
    
    res = set()
    try:
        with open("result.pml", "rb") as fff:
            pml_reader = ProcmonLogsReader(fff)
            while pml_reader:
                row = next(pml_reader)
                full_path = row.path
                dll_name = full_path.split('\\')[-1].lower()
                res.add(dll_name)
    except StopIteration:
        pass
    return list(res)
import re
import pefile
from winreg import CreateKey, EnumValue, HKEY_LOCAL_MACHINE
import os

def check_dynamic(path_pe):
    #global run_now
    #if path_pe == "C:\\windows\\system32\\msdt.exe":
    #    run_now = True
    #    return
    #if run_now == False:
    #    return
    try:
        f = open(path_pe,'rb')
        z = f.read()
    except:
        return 
    x = re.search(b'<autoElevate>true</autoElevate>',z)
    #x = re.search(b'autoElevate',z)
    if x:
        print("[+] " + path_pe)
        list_to_check = export(path_pe)
        print(" - List of dynamic DLL imported: ", end ='')
        print(list_to_check)
        print(" - List of Manifest DLL: ")
        if list_to_check:
            y = re.findall(b'<file[\s\n\r]*loadFrom=.*[\s\n\r]*name=.*[\s\n\r]*\/>', z)
            for i in y:
                k = i.decode('utf-8')
                try:
                    found = re.findall('"(.+?)"', k)[1].lower()
                    print(found, end=' ')
                    list_to_check.remove(found)
                except:
                    pass
        #remove if it in knownDll
            for j in ignore_list:
                try:
                    list_to_check.remove(j)
                except:
                    pass
            print(" - Final result:", end='')
            print(list_to_check)
            if list_to_check:
                global dem 
                dem = dem + 1
                write_excel(path_pe, list_to_check, dem)
                
def check_static(path_pe):
    #readfile
    try:
        f = open(path_pe,'rb')
        z = f.read()
    except:
        return 
    #check Elevate Exe
    x = re.search(b'<autoElevate>true</autoElevate>',z)#x = re.search(b'autoElevate',z)                
    if x:
        pe = pefile.PE(path_pe, fast_load= True)
        pe.parse_data_directories(directories=[1])
        bits = 64 if pe.PE_TYPE == pefile.OPTIONAL_HEADER_MAGIC_PE_PLUS else 32
        print("[+] " + path_pe + ": x" + str(bits))
        list_to_check = list()
        for iid in pe.DIRECTORY_ENTRY_IMPORT:
            dll_name = iid.dll.decode('ascii', errors='replace').lower()
            list_to_check.append(dll_name)
        # remove if it in manifest
        y = re.findall(b'<file[\s\n\r]*loadFrom=.*[\s\n\r]*name=\"[a-z0-9A-Z]*\"[\s\n\r]*\/', z)
        for i in y:
            k = i.decode('utf-8')
            try:
                found = re.findall('"(.+?)"', k)[1]
                list_to_check.remove(found)
            except:
                found = ''
        #remove if it in knownDll
        for j in ignore_list:
            try:
                list_to_check.remove(j)
            except:
                pass
        if list_to_check:
                global dem 
                dem = dem + 1
                write_excel(path_pe, list_to_check, dem)
#check("c:\\windows\\system32\\sysprep\\sysprep.exe")
# def check_dynamic_dir(dirName):
#     try:
#         listOfFile = os.listdir(dirName)
#     except WindowsError:
#         return
#     for entry in listOfFile:
#         fullPath = os.path.join(dirName, entry)
#         if (re.search("^.*\.exe$",entry)):
#             check_dynamic(fullPath)
#         if os.path.isdir(fullPath):
#             check_dir(fullPath)

def check_static_dir(dirName):
    try:
        listOfFile = os.listdir(dirName)
    except WindowsError:
        return
    for entry in listOfFile:
        fullPath = os.path.join(dirName, entry)
        if (re.search("^.*\.exe$",entry)):
            check_static(fullPath)
        if os.path.isdir(fullPath):
            check_static_dir(fullPath)
ignore_list = list()
aKey = "SYSTEM\\CurrentControlSet\\Control\\Session Manager\\KnownDLLs"
aReg = CreateKey(HKEY_LOCAL_MACHINE,aKey)
if aReg:
    i = 0
    while True:
        try:
            bla = EnumValue(aReg,i)
            i= i + 1
        except WindowsError:
            break
        try:
            ignore_list.append(bla[1].lower())
        except:
            pass
dem = 0
run_now = True
ignore_list.append("ntdll.dll")
ignore_list.append("apphelp.dll")
ignore_list.append("comctl32.dll")
count = 0

#if __name__ =='__main__':
check_static_dir("C:\\windows\\system32")
    #check_dynamic_dir("C:\\windows\\system32")
